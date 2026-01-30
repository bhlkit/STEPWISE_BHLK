import numpy as np
import pandas as pd
import statsmodels.api as sm
import customtkinter
import tkinter
import tkinter.filedialog as filedialog
import tkinter.messagebox as messagebox
import matplotlib
import matplotlib.pyplot as plt
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
import os
import traceback
import io
import time
from PIL import Image
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Font, Alignment, PatternFill

# --- KONFIGURASI STYLE ---
matplotlib.use('TkAgg')
customtkinter.set_appearance_mode("Light") 
customtkinter.set_default_color_theme("blue")

# --- PALET WARNA BHLK ---
COLOR_BG_MAIN = "#fcfcfb"       
COLOR_PRIMARY = "#fdb813"       # Gold/Kuning BHLK
COLOR_SECONDARY = "#1f3368"     # Biru Gelap BHLK
COLOR_ACCENT_RED = "#d63031"    
COLOR_WHITE = "#ffffff"         

# Font Style
FONT_MAIN = ("Roboto Medium", 12)
FONT_BOLD = ("Roboto", 13, "bold")
FONT_MONO = ("Consolas", 11)    

# ==========================================================
# CLASS SPLASH SCREEN
# ==========================================================
class SplashScreen(customtkinter.CTk):
    def __init__(self):
        super().__init__()
        
        self.width = 700
        self.height = 354 
        self.overrideredirect(True)
        
        ws = self.winfo_screenwidth()
        hs = self.winfo_screenheight()
        x = (ws/2) - (self.width/2)
        y = (hs/2) - (self.height/2)
        self.geometry('%dx%d+%d+%d' % (self.width, self.height, x, y))
        
        self.configure(fg_color=COLOR_WHITE) 

        current_dir = os.path.dirname(os.path.abspath(__file__))
        image_path = os.path.join(current_dir, "splash_screen.png") 
        
        try:
            pil_image = Image.open(image_path)
            self.splash_img = customtkinter.CTkImage(light_image=pil_image, size=(700, 350))
            self.lbl_img = customtkinter.CTkLabel(self, text="", image=self.splash_img)
            self.lbl_img.pack(side="top", fill="both", expand=False, padx=0, pady=0)
        except Exception as e:
            self.lbl_img = customtkinter.CTkLabel(self, text="Splash Image Not Found", font=FONT_BOLD)
            self.lbl_img.pack(side="top", pady=50)
            print(f"Error loading splash: {e}")

        self.progress = customtkinter.CTkProgressBar(
            self, 
            height=4,                       
            corner_radius=0,                
            progress_color=COLOR_PRIMARY,   
            fg_color=COLOR_WHITE,           
            border_width=0,
            width=700                       
        )
        self.progress.pack(side="bottom", fill="x", padx=0, pady=0)
        self.progress.set(0)

        self.after(50, self.start_loading)

    def start_loading(self):
        steps = [0.1, 0.25, 0.45, 0.60, 0.85, 1.0]
        total_time = 3000 
        step_delay = total_time // len(steps)
        for i, prog in enumerate(steps):
            self.after(i * step_delay, lambda p=prog: self.update_progress(p))
        self.after(total_time + 500, self.finish)

    def update_progress(self, val):
        self.progress.set(val)
        self.update_idletasks()

    def finish(self):
        self.destroy() 


# --- FUNGSI BACA DATA ---
def read_data_native(filename):
    try:
        df = None
        if filename.endswith(('.xlsx', '.xls')):
            df = pd.read_excel(filename, sheet_name=0)
        elif filename.endswith('.csv'):
            df = pd.read_csv(filename)
        elif filename.endswith('.txt'):
            df = pd.read_csv(filename, sep=r'\s+', header=None, skiprows=24, engine='python')
            df.columns = [f'VAR{i}' for i in range(1, len(df.columns) + 1)]
        else:
            raise ValueError("Format file tidak didukung.")

        if df is None or len(df) == 0: raise ValueError("File kosong.")

        df = df.dropna().reset_index(drop=True)
        for col in df.columns:
            df[col] = pd.to_numeric(df[col], errors='coerce')
        df = df.dropna().reset_index(drop=True)
        return df
    except Exception as e:
        messagebox.showerror("Error Baca File", f"{e}")
        return None

def get_descriptive_stats(df):
    return df.describe().T[['count', 'mean', 'std', 'min', 'max']].to_string()

def get_correlation_matrix(df):
    return df.corr().round(4).to_string()

# --- LOGIKA STEPWISE ---
def stepwise_regression_final(df, target_var, fin_threshold):
    included = []
    results = []
    
    while True:
        changed = False
        excluded = list(set(df.columns) - set(included) - {target_var})
        
        best_f = -1.0
        best_feature = None
        best_model = None
        
        for feature in excluded:
            try:
                features_to_try = included + [feature]
                X_poly = sm.add_constant(df[features_to_try])
                y = df[target_var]
                
                model = sm.OLS(y, X_poly).fit()
                f_value = model.fvalue
                if np.isnan(f_value): f_value = 0.0

                if f_value > best_f:
                    best_f = f_value
                    best_feature = feature
                    best_model = model
            except:
                continue

        # LOGIC: Masukkan jika F-Statistic > Threshold
        if best_feature is not None and best_f > fin_threshold:
            included.append(best_feature)
            changed = True
            
            results.append({
                'step': len(results) + 1,
                'action': f"Entered {best_feature}",
                'r_squared': best_model.rsquared,
                'adj_r_squared': best_model.rsquared_adj,
                'aic': best_model.aic,
                'bic': best_model.bic,
                'f_value': best_model.fvalue,
                'f_pvalue': best_model.f_pvalue,
                'model': best_model,
                'scale': best_model.scale
            })
            
        if not changed:
            break
            
    return results

# --- HELPER: RUMUS ---
def generate_equation(model, target_var, use_log):
    params = model.params
    if use_log:
        eq_str = f"Ln({target_var}) = "
    else:
        eq_str = f"{target_var} = "
    
    const_val = params.get('const', 0)
    eq_str += f"{const_val:.4f}"
    
    for var, coef in params.items():
        if var == 'const': continue
        sign = " + " if coef >= 0 else " - "
        val = abs(coef)
        var_name = f"Ln({var})" if use_log else var
        eq_str += f"{sign}{val:.4f}*{var_name}"
    return eq_str

# --- GUI CLASS ---
class StepwiseApp(customtkinter.CTk):
    def __init__(self):
        super().__init__()

        self.title("Aplikasi Stepwise BHLK")
        self.geometry("1280x900")
        self.configure(fg_color=COLOR_BG_MAIN)

        current_dir = os.path.dirname(os.path.abspath(__file__))
        icon_path = os.path.join(current_dir, "bhlk_logo.ico") 
        if os.path.exists(icon_path):
            try:
                self.iconbitmap(icon_path)
            except Exception as e:
                print(f"Gagal load icon bitmap: {e}")
        else:
            print(f"Icon tidak ditemukan di: {current_dir}")

        self.file_path = None
        self.stored_results = None
        self.stored_df_raw = None
        self.stored_df_proc = None
        self.stored_target = None
        self.corr_data_sorted = None 
        self.use_log_status = False
        self.final_equation = "" 
        
        self.grid_columnconfigure(0, weight=1)
        self.grid_rowconfigure(0, weight=0) 
        self.grid_rowconfigure(1, weight=1) 
        self.grid_rowconfigure(2, weight=1) 

        # 1. KONTROL FRAME
        self.ctrl_frame = customtkinter.CTkFrame(self, fg_color=COLOR_WHITE, corner_radius=10, border_width=1, border_color="#e0e0e0")
        self.ctrl_frame.grid(row=0, column=0, padx=20, pady=(20, 10), sticky="ew")
        
        self.btn_file = customtkinter.CTkButton(self.ctrl_frame, text="📂 1. Pilih File Data", command=self.browse_file, font=FONT_BOLD, fg_color=COLOR_SECONDARY, text_color="white", hover_color="#15224a", corner_radius=8, height=35)
        self.btn_file.grid(row=0, column=0, padx=15, pady=15)
        
        self.lbl_file = customtkinter.CTkLabel(self.ctrl_frame, text="File belum dipilih...", text_color="#636e72", font=FONT_MAIN)
        self.lbl_file.grid(row=0, column=1, padx=10, pady=15, sticky="w")

        self.btn_run = customtkinter.CTkButton(self.ctrl_frame, text="▶ JALANKAN ANALISIS", command=self.run_analysis, font=FONT_BOLD, fg_color=COLOR_PRIMARY, text_color=COLOR_SECONDARY, hover_color="#e6a700", corner_radius=8, height=35, width=180)
        self.btn_run.grid(row=0, column=2, padx=15, pady=15)
        
        self.btn_save = customtkinter.CTkButton(self.ctrl_frame, text="💾 SIMPAN EXCEL", command=self.save_to_excel, font=FONT_BOLD, fg_color="transparent", border_width=2, border_color=COLOR_SECONDARY, text_color=COLOR_SECONDARY, hover_color="#eef2ff", corner_radius=8, height=35, state="disabled")
        self.btn_save.grid(row=0, column=3, padx=15, pady=15)

        self.ctrl_frame.grid_columnconfigure(4, weight=1)

        # 2. HASIL TEXT AREA
        self.res_frame = customtkinter.CTkFrame(self, fg_color=COLOR_WHITE, corner_radius=10, border_width=1, border_color="#e0e0e0")
        self.res_frame.grid(row=1, column=0, padx=20, pady=5, sticky="nsew")
        self.res_frame.grid_columnconfigure(0, weight=1)
        self.res_frame.grid_rowconfigure(0, weight=1)
        
        self.txt_out = customtkinter.CTkTextbox(self.res_frame, font=FONT_MONO, text_color=COLOR_SECONDARY, fg_color="#fafafa", corner_radius=5)
        self.txt_out.grid(row=0, column=0, padx=10, pady=10, sticky="nsew")

        # 3. PLOT AREA
        self.plot_container = customtkinter.CTkFrame(self, fg_color="transparent")
        self.plot_container.grid(row=2, column=0, padx=20, pady=10, sticky="nsew")
        self.plot_container.grid_columnconfigure(0, weight=1)
        self.plot_container.grid_rowconfigure(0, weight=1)

        self.tab_view = customtkinter.CTkTabview(self.plot_container, height=300, fg_color=COLOR_WHITE, segmented_button_fg_color="#e0e0e0", segmented_button_selected_color=COLOR_PRIMARY, segmented_button_selected_hover_color="#e6a700", segmented_button_unselected_color="#e0e0e0", segmented_button_unselected_hover_color="#d1d1d1", text_color=COLOR_SECONDARY)
        self.tab_view.grid(row=0, column=0, padx=0, pady=0, sticky="nsew")
        
        self.tab_view.add("Hydrograph")
        self.tab_view.add("Grafik Korelasi")
        self.tab_view.set("Hydrograph")
        
    def browse_file(self):
        f = filedialog.askopenfilename(filetypes=[("Data", "*.xlsx *.csv *.txt")])
        if f:
            self.file_path = f
            self.lbl_file.configure(text=f"File: {os.path.basename(f)}", text_color=COLOR_SECONDARY)

    def run_analysis(self):
        if not self.file_path:
            messagebox.showwarning("Info", "Pilih file dulu.")
            return
        
        # --- PERBAIKAN LOGIC PENTING ---
        # Di Fortran, variabel masuk meskipun F-stat hanya 0.0259.
        # Maka FIN di sini harus 0.0 (atau sangat mendekati 0) agar variabel dipaksa masuk.
        fin = 0.0 
        use_log = False 

        try:
            self.txt_out.delete("0.0", "end")
            self.txt_out.insert("end", "Memproses data... Harap tunggu.\n")
            self.btn_save.configure(state="disabled")
            self.update()

            df = read_data_native(self.file_path)
            if df is None: return
            
            if len(df.columns) < 2:
                messagebox.showerror("Error", "Data harus memiliki minimal 2 kolom (Variabel Bebas & Terikat).")
                return
            target_var = df.columns[-1]

            df_proc = df.copy()

            if use_log:
                for col in df_proc.columns:
                    df_proc[col] = np.log(df_proc[col].replace(0, 1e-5))
            
            corr_matrix = df.corr()
            self.corr_data_sorted = corr_matrix[target_var].drop(target_var).sort_values(ascending=True)

            out = "===========================================================\n"
            out += f" APLIKASI STEPWISE BHLK - REPORT\n"
            out += f" Target (Auto): {target_var}\n"
            out += f" Fin Limit: {fin} (Locked - Force Entry)\n"
            out += "===========================================================\n\n"
            
            out += "--- 1. DESCRIPTIVE STATISTICS ---\n"
            out += get_descriptive_stats(df) + "\n\n"
            out += "--- 2. CORRELATION MATRIX ---\n"
            out += get_correlation_matrix(df) + "\n\n"

            results = stepwise_regression_final(df_proc, target_var, fin_threshold=fin)
            
            if not results:
                out += f"\n[INFO] Tidak ada variabel yang masuk (Fin > {fin}).\n"
                out += "Data mungkin memiliki korelasi yang sangat lemah."
                self.txt_out.delete("0.0", "end")
                self.txt_out.insert("0.0", out)
                return 

            self.stored_results = results
            self.stored_df_raw = df
            self.stored_df_proc = df_proc
            self.stored_target = target_var
            self.use_log_status = use_log
            
            final_model = results[-1]['model']
            self.final_equation = generate_equation(final_model, target_var, use_log)
            self.btn_save.configure(state="normal")

            for res in results:
                m = res['model']
                out += "\n" + "="*80 + "\n"
                out += f"STEP {res['step']}: {res['action']}\n"
                out += "="*80 + "\n"
                
                std_err = np.sqrt(res['scale'])
                mult_r = np.sqrt(res['r_squared'])
                
                out += "MODEL SUMMARY:\n"
                out += f"{'Statistic':<25} {'Value':<15}\n"
                out += "-"*40 + "\n"
                out += f"{'Std Error Estimate':<25} {std_err:.4f}\n"
                out += f"{'Multiple Correlation':<25} {mult_r:.4f}\n"
                out += f"{'R-squared':<25} {res['r_squared']:.4f}\n"
                out += f"{'Adj. R-squared':<25} {res['adj_r_squared']:.4f}\n"
                out += f"{'F-statistic':<25} {res['f_value']:.4f}\n"
                out += f"{'AIC':<25} {res['aic']:.2f}\n"
                out += "-"*40 + "\n\n"
                
                betas = m.params / m.model.endog.std()

                cdf = pd.DataFrame({
                    'Coef': m.params,
                    'StdErr': m.bse,
                    't': m.tvalues,
                    'PVal': m.pvalues,
                    'Beta': betas 
                })
                
                out += "MODEL COEFFICIENTS:\n"
                out += f"{'Variable':<20} {'Coef':<10} {'StdErr':<10} {'Beta':<10} {'t-Stat':<8} {'P-Val':<8}\n"
                out += "-"*75 + "\n"
                for idx, row in cdf.iterrows():
                    beta_val = f"{row['Beta']:.4f}"
                    out += f"{str(idx):<20} {row['Coef']:<10.4f} {row['StdErr']:<10.4f} {beta_val:<10} {row['t']:<8.2f} {row['PVal']:<8.4f}\n"

            out += "\n" + "="*80 + "\n"
            out += " FINAL MODEL SUMMARY & EQUATION\n"
            out += "="*80 + "\n"
            out += "RUMUS REGRESI (MODEL MATH):\n"
            out += f"{self.final_equation}\n\n"
            
            out += f"{'Step':<5} {'Action':<25} {'F-Val':<10} {'R2':<10}\n"
            out += "-"*80 + "\n"
            for r in results:
                out += f"{r['step']:<5} {r['action']:<25} {r['f_value']:<10.4f} {r['r_squared']:<10.4f}\n"

            y_pred_proc = final_model.fittedvalues
            y_act_proc = df_proc[target_var]
            
            if use_log:
                y_pred_real = np.exp(y_pred_proc)
                y_act_real = np.exp(y_act_proc)
            else:
                y_pred_real = y_pred_proc
                y_act_real = y_act_proc
            
            out += "\n" + "="*80 + "\n"
            out += " PREDICTION TABLE (First 50 Data Points)\n"
            out += "-"*80 + "\n"
            
            limit = min(len(y_act_real), 50)
            for i in range(limit):
                obs = y_act_real.iloc[i]
                pred = y_pred_real.iloc[i]
                diff = obs - pred
                out += f"{i+1:<5} {obs:<15.3f} {pred:<15.3f} {diff:<10.3f}\n"
            
            self.txt_out.delete("0.0", "end")
            self.txt_out.insert("0.0", out)
            
            self.plot_hydrograph(y_act_real, y_pred_real)
            self.plot_correlation_graph()

        except Exception as e:
            err_msg = traceback.format_exc()
            messagebox.showerror("Critical Error", f"{e}\n\n{err_msg}")
            print(err_msg)

    def plot_hydrograph(self, y_act, y_pred):
        for w in self.tab_view.tab("Hydrograph").winfo_children(): w.destroy()
        try:
            fig, ax = plt.subplots(figsize=(10, 3.5), dpi=100, facecolor=COLOR_WHITE)
            ax.set_facecolor(COLOR_WHITE)
            
            ax.plot(y_act.values.flatten(), label='Observasi', color=COLOR_SECONDARY, linewidth=1.5)
            ax.plot(y_pred.values.flatten(), label='Prediksi', color=COLOR_ACCENT_RED, linestyle='--', linewidth=1.5)
            
            ax.set_title("Hydrograph Comparison", color=COLOR_SECONDARY, fontsize=10, fontweight='bold')
            ax.set_ylabel("Debit", color=COLOR_SECONDARY, fontsize=9)
            ax.tick_params(colors=COLOR_SECONDARY, labelsize=8)
            
            legend = ax.legend(fontsize=8)
            plt.setp(legend.get_texts(), color=COLOR_SECONDARY)
            ax.grid(True, alpha=0.2, color=COLOR_SECONDARY)
            
            for spine in ax.spines.values():
                spine.set_edgecolor(COLOR_SECONDARY)
                spine.set_alpha(0.5)

            plt.tight_layout()
            
            canvas = FigureCanvasTkAgg(fig, master=self.tab_view.tab("Hydrograph"))
            canvas.draw()
            canvas.get_tk_widget().pack(fill="both", expand=True)
        except Exception as e:
            print(f"Plot Error: {e}")

    def plot_correlation_graph(self):
        for w in self.tab_view.tab("Grafik Korelasi").winfo_children(): w.destroy()
        if self.corr_data_sorted is None: return

        try:
            fig, ax = plt.subplots(figsize=(10, 3.5), dpi=100, facecolor=COLOR_WHITE)
            ax.set_facecolor(COLOR_WHITE)
            
            x_labels = self.corr_data_sorted.index.tolist()
            y_values = self.corr_data_sorted.values.tolist()
            x_pos = range(len(x_labels))

            ax.plot(x_pos, y_values, color=COLOR_SECONDARY, linewidth=2, zorder=1)
            ax.scatter(x_pos, y_values, color=COLOR_PRIMARY, s=60, zorder=2, edgecolors=COLOR_SECONDARY)

            ax.set_xticks(x_pos)
            ax.set_xticklabels(x_labels, rotation=45, ha='right', fontsize=9)
            ax.set_title(f"Grafik Korelasi (r) Tiap Stasiun terhadap {self.stored_target}", color=COLOR_SECONDARY, fontsize=10, fontweight='bold')
            ax.set_ylabel("Nilai Korelasi (r)", color=COLOR_SECONDARY, fontsize=9)
            ax.tick_params(colors=COLOR_SECONDARY, labelsize=8)
            ax.grid(True, linestyle='--', alpha=0.2, color=COLOR_SECONDARY)

            for spine in ax.spines.values():
                spine.set_edgecolor(COLOR_SECONDARY)
                spine.set_alpha(0.5)

            for i, v in enumerate(y_values):
                ax.text(i, v + 0.02, f"{v:.2f}", ha='center', fontsize=8, color=COLOR_SECONDARY)

            plt.tight_layout()
            
            canvas = FigureCanvasTkAgg(fig, master=self.tab_view.tab("Grafik Korelasi"))
            canvas.draw()
            canvas.get_tk_widget().pack(fill="both", expand=True)
        except Exception as e:
            print(f"Plot Error Corr: {e}")

    def save_to_excel(self):
        if not self.stored_results:
            messagebox.showwarning("Warning", "Belum ada hasil analisis.")
            return
            
        filename = filedialog.asksaveasfilename(defaultextension=".xlsx", 
                                                filetypes=[("Excel Files", "*.xlsx")])
        if not filename: return
        
        try:
            with pd.ExcelWriter(filename, engine='openpyxl') as writer:
                # --- A. SHEET: LAPORAN STEPWISE ---
                pd.DataFrame().to_excel(writer, sheet_name='Laporan Stepwise')
                workbook = writer.book
                ws_step = writer.sheets['Laporan Stepwise']
                bold_font = Font(bold=True)
                current_row = 1
                
                for res in self.stored_results:
                    ws_step.cell(row=current_row, column=1, value=f"STEP {res['step']}: {res['action']}")
                    ws_step.cell(row=current_row, column=1).font = Font(bold=True, size=12, color="0000FF")
                    current_row += 2
                    
                    ws_step.cell(row=current_row, column=1, value="Model Summary")
                    ws_step.cell(row=current_row, column=1).font = bold_font
                    current_row += 1
                    
                    summary_data = {
                        'Statistic': ['Std Error Estimate', 'Multiple Correlation (R)', 'R-squared', 'Adj. R-squared', 'F-statistic', 'AIC'],
                        'Value': [np.sqrt(res['scale']), np.sqrt(res['r_squared']), res['r_squared'], res['adj_r_squared'], res['f_value'], res['aic']]
                    }
                    pd.DataFrame(summary_data).to_excel(writer, sheet_name='Laporan Stepwise', startrow=current_row-1, index=False)
                    current_row += len(summary_data['Statistic']) + 3
                    
                    ws_step.cell(row=current_row, column=1, value="Model Coefficients")
                    ws_step.cell(row=current_row, column=1).font = bold_font
                    current_row += 1
                    
                    m = res['model']
                    betas = m.params / m.model.endog.std()

                    coef_df = pd.DataFrame({
                        'Variable': m.params.index,
                        'Coef': m.params.values,
                        'Std Error': m.bse.values,
                        'Beta': betas.values, 
                        't-Stat': m.tvalues.values,
                        'P-Value': m.pvalues.values
                    })
                    coef_df.to_excel(writer, sheet_name='Laporan Stepwise', startrow=current_row-1, index=False)
                    current_row += len(coef_df) + 4
                
                ws_step.column_dimensions['A'].width = 25
                ws_step.column_dimensions['B'].width = 15
                
                # --- B. SHEET: PREDIKSI & GRAFIK ---
                final_model = self.stored_results[-1]['model']
                y_pred_proc = final_model.fittedvalues
                y_act_proc = self.stored_df_proc[self.stored_target]
                
                if self.use_log_status:
                    y_pred_real = np.exp(y_pred_proc)
                    y_act_real = np.exp(y_act_proc)
                else:
                    y_pred_real = y_pred_proc
                    y_act_real = y_act_proc

                ws_pred = workbook.create_sheet("Prediksi & Grafik")
                writer.sheets["Prediksi & Grafik"] = ws_pred
                
                ws_pred.cell(row=1, column=1, value="RUMUS REGRESI:")
                ws_pred.cell(row=1, column=1).font = bold_font
                ws_pred.cell(row=2, column=1, value=self.final_equation)
                ws_pred.cell(row=2, column=1).font = Font(color="FF0000", size=11)
                
                pred_df = pd.DataFrame({
                    'Observasi': y_act_real.values.flatten(),
                    'Prediksi': y_pred_real.values.flatten(),
                    'Selisih (Diff)': (y_act_real - y_pred_real).values.flatten()
                })
                pred_df.to_excel(writer, sheet_name='Prediksi & Grafik', startrow=4, index_label='No')
                
                # --- C. SHEET: STATISTIK & KORELASI ---
                self.stored_df_raw.describe().T.to_excel(writer, sheet_name='Statistik Deskriptif')
                self.stored_df_raw.corr().to_excel(writer, sheet_name='Matriks Korelasi')
                ws_corr = writer.sheets['Matriks Korelasi']

                # --- D. MENYISIPKAN GAMBAR ---
                # 1. Hydrograph
                try:
                    plt.figure(figsize=(10, 4))
                    plt.plot(y_act_real.values.flatten(), label='Observasi', color='blue', linewidth=1)
                    plt.plot(y_pred_real.values.flatten(), label='Prediksi', color='red', linestyle='--', linewidth=1)
                    plt.title("Hydrograph Comparison")
                    plt.legend()
                    plt.grid(True, alpha=0.3)
                    
                    img_buffer = io.BytesIO()
                    plt.savefig(img_buffer, format='png')
                    img_buffer.seek(0)
                    plt.close('all') 
                    
                    xl_img = XLImage(img_buffer)
                    ws_pred.add_image(xl_img, 'F5')
                except Exception as img_err:
                    print(f"Gagal menyisipkan grafik hydro: {img_err}")

                # 2. Grafik Korelasi (PERBAIKAN LOGIC DISINI)
                if self.corr_data_sorted is not None:
                    try:
                        plt.figure(figsize=(10, 5))
                        x_labels = self.corr_data_sorted.index.tolist()
                        y_values = self.corr_data_sorted.values.tolist()
                        x_pos = range(len(x_labels))

                        plt.plot(x_pos, y_values, color='#4472C4', linewidth=2, zorder=1)
                        plt.scatter(x_pos, y_values, color='red', s=50, zorder=2)
                        
                        plt.xticks(x_pos, x_labels, rotation=45, ha='right')
                        plt.title(f"Grafik Korelasi (r) Tiap Stasiun terhadap {self.stored_target}", color=COLOR_SECONDARY, fontsize=10, fontweight='bold')
                        plt.ylabel("Nilai Korelasi (r)")
                        plt.grid(True, linestyle='--', alpha=0.5)
                        plt.tight_layout()

                        # Simpan ke buffer utama
                        img_buffer_c = io.BytesIO()
                        plt.savefig(img_buffer_c, format='png')
                        plt.close('all')
                        
                        # Reset pointer buffer utama
                        img_buffer_c.seek(0) 
                        
                        # Simpan ke Sheet Matriks Korelasi (Gunakan buffer utama)
                        xl_img_c = XLImage(img_buffer_c)
                        last_row = len(self.stored_df_raw.columns) + 5
                        ws_corr.add_image(xl_img_c, f'A{last_row}')

                        # Simpan ke Sheet Prediksi & Grafik (BUAT BUFFER BARU / COPY)
                        # Ini solusi untuk error "closed file": Buat BytesIO terpisah berisi data yang sama
                        img_buffer_c2 = io.BytesIO(img_buffer_c.getvalue())
                        xl_img_c_pred = XLImage(img_buffer_c2)
                        ws_pred.add_image(xl_img_c_pred, 'F27')

                    except Exception as img_err:
                        print(f"Gagal menyisipkan grafik korelasi: {img_err}")

            messagebox.showinfo("Sukses", f"Laporan Excel Lengkap tersimpan di:\n{filename}")
            
        except Exception as e:
            err = traceback.format_exc()
            messagebox.showerror("Gagal Menyimpan", f"Error saving excel:\n{e}\n{err}")

if __name__ == "__main__":
    splash = SplashScreen()
    splash.mainloop()
    
    app = StepwiseApp()
    app.mainloop()